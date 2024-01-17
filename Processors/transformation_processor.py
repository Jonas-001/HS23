import ifcopenshell
import openpyxl
from ifcopenshell.api import run
import os

def transform_elements(target_folder, excel_path, ifc_filename):  # Füge das dritte Argument hinzu
    # Benutzereingabe für Basisordner
    source_folder = target_folder  # Änderung: Verwende den Zielordner als Basisordner
    
    # Überprüfe, ob der Zielordner existiert
    if not os.path.exists(source_folder):
        print(f"Der Zielordner {source_folder} existiert nicht.")
        return

    # Erstellen des vollständigen Pfads zur Excel-Datei im Basisordner
    xls_file_path = excel_path

    # Überprüfe, ob die IFC-Datei im Zielordner vorhanden ist
    ifc_filename_with_extension = ifc_filename + ".ifc"  # Füge die .ifc-Erweiterung hinzu
    ifc_file_path = os.path.join(source_folder, ifc_filename_with_extension)

    if not os.path.exists(ifc_file_path):
        print(f"Die IFC-Datei {ifc_file_path} wurde nicht gefunden.")
        return

    print(f"IFC-Datei gefunden: {ifc_file_path}")

    # Öffnen der IFC-Datei
    ifc_file = ifcopenshell.open(ifc_file_path)

    # Öffnen der Excel-Datei
    workbook = openpyxl.load_workbook(xls_file_path)

    # Iteration durch alle Sheets in der Excel-Datei
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        # Iteration durch die Zeilen in der Excel-Datei
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            guid = row[0]  # Die GUID-ID befindet sich in der ersten Spalte
            new_name = row[6]  # Der neue Name befindet sich in der siebten Spalte
            pset_name = row[7]  # Der Pset-Name befindet sich in der achten Spalte
            property_value = row[8]  # Der Wert aus der neunten Spalte
            property_info = "Material"  # Der festgelegte Property-Name

            # Versuch, das entsprechende IFC-Element anhand der GUID-ID zu finden
            try:
                ifc_element = ifc_file.by_guid(guid)
            except RuntimeError:
                print(f"Die GUID {guid} wurde in der IFC-Datei nicht gefunden. Überspringe diese Zeile.")
                continue  # Springe zur nächsten Zeile in der Excel-Datei

            # Überprüfen, ob das IFC-Element gefunden wurde
            if ifc_element is not None:
                # Überprüfen des IFC-Elementtyps
                if ifc_element.is_a("IfcWall"):
                    element_type = "Wall"
                elif ifc_element.is_a("IfcSlab"):
                    element_type = "Slab"
                elif ifc_element.is_a("IfcBeam"):
                    element_type = "Beam"
                elif ifc_element.is_a("IfcColumn"):
                    element_type = "Column"
                else:
                    element_type = "Unbekannter Typ"

                # Aktualisieren des Namens des IFC-Elements
                ifc_element.Name = new_name

                # Hinzufügen des Pset und der Property
                pset = run("pset.add_pset", ifc_file, product=ifc_element, name=pset_name)
                run("pset.edit_pset", ifc_file, pset=pset, properties={property_info: property_value})

                print(f"Element mit GUID {guid} gefunden. Typ: {element_type}, Neuer Name: {new_name}, Pset: {pset_name}, Property: {property_info}")
            else:
                print(f"Die GUID {guid} wurde in der IFC-Datei nicht gefunden. Überspringe diese Zeile.")

    # Speichern der Änderungen in der IFC-Datei
    ifc_file.write(ifc_file_path)

    # Schließen der Excel-Datei
    workbook.close()

    print("Die IFC-Datei wurde aktualisiert und gespeichert.")