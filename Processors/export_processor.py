import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
import random
import Helpers.Data_Helpers as hlp

def export_data(export_lst, conf_lst, export_folder):

    if len(export_lst) > 0:

        xls_file_name = "Export_Daten_IFC.xlsx"
        exp_path = os.path.join(export_folder, xls_file_name)
        
        check_file = os.path.isfile(exp_path)

        if not check_file:
            pd.DataFrame([]).to_excel(exp_path)

        for exp_set in export_lst:

            col_lst = ["GUID", "Kategorie", "Materialinformationen", "Storey", "Building"]     
            
            _target_obj = hlp.filterConfig(exp_set.branch, conf_lst)
            print(_target_obj)
                        
            if _target_obj is not None:
                for _p_l in _target_obj.prop_list:
                    _ps = _p_l.pset
                    _pr = _p_l.prop

                    if str(_ps) == "nan":
                        _title = "{}".format(_pr)
                    else:
                        _title = "{}:{}".format(_ps, _pr)

                    col_lst.append(_title)
                    print(_title)

                df = pd.DataFrame(exp_set.data, columns=col_lst)

                # Extrahiere Storey- und Gebäudeinformationen und füge sie zur Excel-Tabelle hinzu
                storey_col = df["Storey"].apply(lambda x: x if x else "")
                building_col = df["Building"].apply(lambda x: x if x else "")
                df["Storey"] = storey_col
                df["Building"] = building_col

                # Extrahiere Materialinformationen und füge sie zur Excel-Tabelle hinzu
                material_info_col = df["Materialinformationen"].apply(lambda x: x[0] if x else None)
                df["Materialinformationen"] = material_info_col.apply(lambda x: str(x) if x else "")
                
                # Excel-Datei schreiben
                with pd.ExcelWriter(exp_path, engine='openpyxl', if_sheet_exists='replace', mode='a') as writer:
                    # Schreibe den DataFrame in die Excel-Datei
                    df.to_excel(writer, sheet_name=exp_set.branch, index=False, header=False)

                    # Lade das Arbeitsblatt, um benutzerdefinierte Formatierungen hinzuzufügen
                    sheet = writer.sheets[exp_set.branch]
                
                    # Styling-Funktionen
                    def style_df(val):
                        return [
                            f'border: 2px solid white; background-color: rgba(0,128,0,0.3)' if i % 2 == 0 else
                            f'border: 2px solid white; background-color: rgba(0,128,0,0.1)'
                            for i, _ in enumerate(val)
                        ]

                    def highlight_header(val):
                        return [
                            f'font-weight: bold; background-color: rgba(0,128,0,0.6)' if i == 0 else ''
                            for i, _ in enumerate(val)
                        ]

                    # Initialisiere die Zuordnung von Werten zu Rahmenfarben
                    border_color_mapping = {}

                    # Wende benutzerdefinierte Formatierungen auf das Arbeitsblatt an
                    for idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=df.shape[1], max_row=df.shape[0] + 1), 2):
                        for col_idx, cell in enumerate(row, 1):
                            # Rahmen für Spalte 2 und 3
                            if col_idx in [2, 3]:
                                if cell.value:
                                    cell_border_color = 'FFFFFF'  # Standard-Rahmenfarbe
                                    if cell.value not in border_color_mapping:
                                        # Wenn der Wert nicht in der Zuordnung ist, generiere eine zufällige Farbe
                                        border_color_mapping[cell.value] = f'FF{random.randint(0, 0xFFFFFF):06x}'
                                    cell_border_color = border_color_mapping[cell.value]
                                    # Rahmenformatierung für nicht leere Zellen
                                    cell.border = Border(
                                        left=Side(border_style='medium', color=cell_border_color),
                                        right=Side(border_style='medium', color=cell_border_color),
                                        top=Side(border_style='medium', color=cell_border_color),
                                        bottom=Side(border_style='medium', color=cell_border_color)
                                    )
                                    # Andere Formatierungen für die Spalten 2 und 15
                                    if col_idx == 2 or col_idx == 15:
                                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                            # Hintergrundfarben und Header-Styling für den Rest des Arbeitsblatts bleiben unverändert

                            # Rahmen für die letzte Spalte
                            elif col_idx == df.shape[1]:
                                if cell.value:
                                    cell_border_color = 'FFFFFF'  # Standard-Rahmenfarbe
                                    if cell.value not in border_color_mapping:
                                        # Wenn der Wert nicht in der Zuordnung ist, generiere eine zufällige Farbe
                                        border_color_mapping[cell.value] = f'FF{random.randint(0, 0xFFFFFF):06x}'
                                    cell_border_color = border_color_mapping[cell.value]
                                    # Rahmenformatierung für nicht leere Zellen
                                    cell.border = Border(
                                        left=Side(border_style='medium', color=cell_border_color),
                                        right=Side(border_style='medium', color=cell_border_color),
                                        top=Side(border_style='medium', color=cell_border_color),
                                        bottom=Side(border_style='medium', color=cell_border_color)
                                    )

                    # Hier wird die `round`-Funktion verwendet, um die Zahlen auf zwei Stellen nach dem Komma zu runden,
                    # aber nur, wenn es sich um einen numerischen Typ handelt
                    for row in sheet.iter_rows(min_row=2, max_col=df.shape[1] - 1, max_row=df.shape[0] + 1):
                        for col_idx, cell in enumerate(row, 1):
                            if isinstance(cell.value, (int, float)):
                                cell.value = round(cell.value, 2)

                    # Wende Header-Styling an
                    for cell in sheet.iter_rows(min_row=1, max_row=1, max_col=df.shape[1]):
                        for col_idx, header_cell in enumerate(cell, 1):
                            header_cell.font = Font(bold=True)
                            header_cell.fill = PatternFill(start_color='A8E6AE', end_color='A8E6AE', fill_type='solid', fgColor="A8E6AEFF")  # Undurchsichtige Farbe
                            header_cell.value = df.columns[col_idx - 1]

                    # Abwechselnde Hintergrundfarben für die restlichen Zeilen
                    for row_idx in range(2, df.shape[0] + 2):
                        if row_idx % 2 == 0:
                            # Gerade Zeilen: RGB 1
                            fill_color = 'DBF5DD'  # RGB 1
                        else:
                            # Ungerade Zeilen: RGB 2
                            fill_color = 'EFFBF0'  # RGB 2
                        
                        for cell in sheet.iter_rows(min_row=row_idx, max_row=row_idx, max_col=df.shape[1]):
                            for col_idx, data_cell in enumerate(cell, 1):
                                data_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

                    # Festlegung eines zusätzlichen Breitenwertes
                    zusätzliche_breite = 4

                    # Automatische Anpassung der Spaltenbreite
                    for column in sheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:  # Erforderlich, um Fehler bei leeren Zellen zu vermeiden
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + zusätzliche_breite)
                        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        print("Final Step: Finished Export Excel...")
