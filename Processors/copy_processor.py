import shutil
from openpyxl import load_workbook
import ifcopenshell
import os

def keep_elements_from_excel(ifc_path, excel_path, target_folder, filename):
    # Erstellen des vollständigen Pfads zur Excel-Datei im Basisordner
    xls_file_path = excel_path

    # Erstellen des vollständigen Pfads zur IFC-Datei im Basisordner
    ifc_file_path = ifc_path

    # Öffnen der Original-IFC-Datei
    original_ifc_file = ifcopenshell.open(ifc_file_path)

    # Kopieren der Original-IFC-Datei
    target_ifc_filename = filename + ".ifc"
    copy_ifc_file_path = os.path.join(target_folder, target_ifc_filename)
    shutil.copyfile(ifc_file_path, copy_ifc_file_path)
    print(f"IFC-Datei wurde kopiert und unter {copy_ifc_file_path} gespeichert.")
    
    # Öffnen der kopierten IFC-Datei
    copy_ifc_file = ifcopenshell.open(copy_ifc_file_path)

    # Öffnen der Excel-Datei
    workbook = load_workbook(xls_file_path)

    # Liste zum Speichern der Elemente, die beibehalten werden sollen
    elements_to_keep = []

    # Pauschal alle benötigten Instanzen und Referenzen in die Liste elements_to_keep aufnehmen
    for element_type in ["IfcSite", "IfcBuilding", "IfcBuildingStorey", "IfcOwnerHistory"]:
        for element in copy_ifc_file.by_type(element_type):
            elements_to_keep.append(element)

    # Iteration durch alle Sheets in der Excel-Datei
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        # Iteration durch die Zeilen in der Excel-Datei
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            guid = str(row[0]).strip()  # Die GUID-ID befindet sich in der ersten Spalte, Leerzeichen entfernen
            x_value = str(row[9]).strip()  # Wert in Spalte 10 (0-basiert) extrahieren

            # Überprüfe, ob die GUID in der IFC-Datei vorhanden ist und der Wert in Spalte 10 ein "X" ist
            found = False
            if x_value.upper() == "X":
                for element in copy_ifc_file.by_type("IfcProduct"):
                    if element.GlobalId.strip() == guid:
                        found = True
                        elements_to_keep.append(element)  # Element zur Liste hinzufügen
                        break

                # Falls gefunden, Ausgabe und weiter zur nächsten GUID
                if found:
                    print(f"Das Element mit der GUID {guid} wurde behalten.")
                else:
                    print(f"Die GUID {guid} wurde nicht in der IFC-Datei gefunden.")
            else:
                print(f"Die GUID {guid} wurde aufgrund des Werts in Spalte 10 nicht berücksichtigt.")

    # Durchgehen aller Elemente und löschen, falls die GUID nicht übereinstimmt
    for element in copy_ifc_file.by_type("IfcProduct"):
        if element not in elements_to_keep:
            copy_ifc_file.remove(element)

    # Speichern der veränderten IFC-Datei
    copy_ifc_file.write(copy_ifc_file_path)

    print("IFC-Datei wurde erfolgreich bearbeitet und unter dem angegebenen Namen gespeichert.")

    # Rückgabe des Dateinamens der kopierten IFC-Datei
    return target_ifc_filename
