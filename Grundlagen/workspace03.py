import ifcopenshell
import openpyxl

# Öffnen die IFC-Datei
ifc_file = ifcopenshell.open(r"C:\Users\joni_\Documents\Studium HSLU\VS_Projects\IFC-Editor\ARC_Modell_NEST_230328.ifc")

# Öffnen die Excel-Datei mit dem Regelwerk
workbook = openpyxl.load_workbook(r"C:\Users\joni_\Documents\Studium HSLU\VS_Projects\IFC-Editor\Test.xslx)
worksheet = workbook.active

# Erstelle eine neue Excel-Datei oder aktualisieren Sie eine vorhandene
ergebnis_workbook = openpyxl.Workbook()
ergebnis_worksheet = ergebnis_workbook.active

# Iterieren Sie über die Zeilen im Regelwerk
for row in worksheet.iter_rows(min_row=2, values_only=True):  # Starte bei Zeile 2, falls die erste Zeile Überschriften enthält
    regel_name = row[0]  # Annahme: Der Name der Regel befindet sich in der ersten Spalte
    regel_bedingung = row[1]  # Annahme: Die Bedingung der Regel befindet sich in der zweiten Spalte

    # Iterieren Sie über die Elemente in der IFC-Datei und wenden Sie das Regelwerk an
    for ifc_element in ifc_file.by_type('IFCElement'):
        # Überprüfen Sie, ob die Bedingung der Regel erfüllt ist
        if bedingung_erfuellt(ifc_element, regel_bedingung):
            # Extrahieren Sie die erforderlichen Daten aus dem IFC-Element
            daten = extrahiere_daten(ifc_element)

            # Schreibe die Daten in die Ergebnis-Excel-Datei
            ergebnis_worksheet.append([regel_name] + daten)

# Speichere die Ergebnisse in einer neuen Excel-Datei
ergebnis_workbook.save(r"C:\Users\joni_\Documents\Studium HSLU\VS_Projects\IFC-Editor\ARC_Modell_NEST_230328.ifc")

# Schliesse die IFC-Datei
ifc_file.close()
