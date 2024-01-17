""" import ifcopenshell
import ifcopenshell.util.element
from ifcopenshell.util.selector import Selector
import os
import pandas as pd
import Helpers.Data_Helpers as hlp

# Globale Variable - Liste für exportierte Daten
export_lst = []

# Funktion zur Verarbeitung der IFC-Daten
def process_ifc_data(conf_lst):
    # Iteriere durch jede Konfiguration in der Liste
    for _c in conf_lst:
        # Erstelle den vollständigen Pfad zur IFC-Datei
        _source_path = os.path.join(_c.folder, _c.source)
        
        # Überprüfe, ob die Datei existiert
        if not os.path.isfile(_source_path):
            print(f"Die Datei {_source_path} existiert nicht.")
            continue

        # Öffne die IFC-Datei
        ifc_file = ifcopenshell.open(_source_path)
        
        # Extrahiere Elemente vom angegebenen IFC-Typ
        elements = ifc_file.by_type(_c.category)
        
        # Liste zur temporären Speicherung der exportierten Daten für die aktuelle Konfiguration
        tmp_exp_data_from_current_config = []

        # Iteriere durch jedes IFC-Element des angegebenen Typs
        for _e in elements:
            # Liste für temporäre Speicherung der Daten eines einzelnen IFC-Elements
            tmp_exp_data = []

            # Extrahiere allgemeine Informationen über das IFC-Element
            act_info = _e.get_info()

            try:
                # Füge die GlobalId und den Typ des IFC-Elements zur temporären Liste hinzu
                tmp_exp_data.append(act_info.GlobalId)
                tmp_exp_data.append(act_info.type)
            except Exception as e:
                # Behandle den Fehler, wenn Informationen nicht abgerufen werden können
                print(f"Fehler beim Abrufen von Informationen: {e}")
                continue

            # Extrahiere die Pset-Eigenschaften des IFC-Elements
            act_prop = ifcopenshell.util.element.get_psets(_e, psets_only=False)

            # Iteriere durch jede konfigurierte Eigenschaft
            for _p in _c.prop_list:
                try:
                    # Überprüfe, ob ein Pset für die Eigenschaft konfiguriert ist
                    if pd.notna(_p.pset):
                        # Extrahiere den Wert aus dem Pset des IFC-Elements
                        act_val = act_prop.get(_p.pset, {}).get(_p.prop)
                    else:
                        # Extrahiere den Wert aus den allgemeinen Eigenschaften des IFC-Elements
                        act_val = act_info.get(_p.prop)

                    # Füge den Wert zur temporären Liste hinzu
                    tmp_exp_data.append(act_val)
                except Exception as e:
                    # Behandle den Fehler, wenn der Wert nicht abgerufen werden kann
                    print(f"Fehler beim Abrufen von Eigenschaften: {e}")
                    tmp_exp_data.append(None)

            # Füge die Daten des aktuellen IFC-Elements zur temporären Liste hinzu
            tmp_exp_data_from_current_config.append(tmp_exp_data)

            # Rufe die Funktion get_ifc_materials auf, um Materialinformationen zu extrahieren
            materials_info = get_ifc_materials(_e)
            # materials_info ist eine Liste von Dictionaries mit Materialinformationen
            # Du kannst diese Informationen nach Bedarf weiterverarbeiten oder speichern

        # Erstelle einen Export-Holder für die aktuelle Konfiguration
        exp_holder = hlp.Exp_Holder(_c.branch, tmp_exp_data_from_current_config)
        # Füge den Export-Holder zur globalen Liste hinzu
        export_lst.append(exp_holder)

    # Gib die gesamte Liste der exportierten Daten zurück
    return export_lst

# Funktion zur Extraktion von IFC-Materialinformationen
def get_ifc_materials(ifc_product):
    # Liste für Materialinformationen
    material_list = []

    # Überprüfe, ob das IFC-Produkt vorhanden ist
    if ifc_product:
        # Extrahiere das Material des IFC-Produkts
        ifc_material = ifcopenshell.util.element.get_material(ifc_product)
        # Überprüfe, ob das Material vorhanden ist
        if ifc_material:    
            # Fall: Material ist ein einfaches IFC-Material
            if ifc_material.is_a('IfcMaterial'):
                mat_dict = {}
                mat_dict["Material"] = ifc_material.Name
                mat_dict["LayerThickness"] = ""  # Platz für zusätzliche Informationen, hier leer gelassen                   
                material_list.append(mat_dict)

            # Fall: Material ist eine Liste von Materialien
            if ifc_material.is_a('IfcMaterialList'):
                for materials in ifc_material.Materials:
                    mat_dict = {}
                    mat_dict["Material"] = materials.Name
                    mat_dict["LayerThickness"] = ""  # Platz für zusätzliche Informationen, hier leer gelassen       
                    material_list.append(mat_dict)

            # Fall: Material besteht aus Materialbestandteilen
            if ifc_material.is_a('IfcMaterialConstituentSet'):
                for material_constituents in ifc_material.MaterialConstituents:
                    mat_dict = {}
                    mat_dict["Material"] = material_constituents.Material.Name
                    mat_dict["LayerThickness"] = ""  # Platz für zusätzliche Informationen, hier leer gelassen  
                    material_list.append(mat_dict)

            # Fall: Material besteht aus Materiallagen
            if ifc_material.is_a('IfcMaterialLayerSetUsage'):
                for material_layer in ifc_material.ForLayerSet.MaterialLayers:
                    mat_dict = {}
                    mat_dict["Material"] = material_layer.Material.Name
                    mat_dict["LayerThickness"] = material_layer.LayerThickness 
                    material_list.append(mat_dict)

            # Fall: Material besteht aus Materialprofilen
            if ifc_material.is_a('IfcMaterialProfileSetUsage'):
                for material_profile in (ifc_material.ForProfileSet.MaterialProfiles):
                    mat_dict = {}
                    mat_dict["Material"] = material_profile.Material.Name
                    mat_dict["LayerThickness"] = ""  # Platz für zusätzliche Informationen, hier leer gelassen  
                    material_list.append(mat_dict)

    # Wenn keine Materialinformationen gefunden wurden, füge None zur Liste hinzu
    if not material_list:
        material_list.append(None)

    # Gib die Liste der Materialinformationen zurück
    return material_list """

import ifcopenshell
import openpyxl

# Dateipfade zu IFC- und Excel-Dateien
ifc_file_path = r"C:\Users\joni_\Documents\Studium HSLU\VS_Projects\IFC-Editor\ARC_Modell_NEST_230328.ifc"
excel_file_path = r"C:\Users\joni_\Documents\Studium HSLU\VS_Projects\IFC-Editor\Export_Daten_IFC - Test.xlsx"

# Öffnen der IFC-Datei
ifc_file = ifcopenshell.open(ifc_file_path)

# Öffnen der Excel-Datei
workbook = openpyxl.load_workbook(excel_file_path)

# Iteration durch alle Sheets in der Excel-Datei
for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]

    # Iteration durch die Zeilen in der Excel-Datei
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        guid = row[0]  # Die GUID-ID befindet sich in der ersten Spalte
        new_name = row[6]  # Der neue Name befindet sich in der siebten Spalte

        # Suchen des entsprechenden IFC-Elements anhand der GUID-ID
        ifc_element = ifc_file.by_guid(guid)

        # Überprüfen, ob das IFC-Element gefunden wurde
        if ifc_element is not None:
            # Überprüfen des IFC-Elementtyps
            if ifc_element.is_a("IfcWall"):
                element_type = "Wall"
            elif ifc_element.is_a("IfcSlab"):
                element_type = "Slab"
            elif ifc_element.is_a("IfcBeam"):
                element_type = "Beam"
            elif ifc_element.is_a("IfcColumn"):
                element_type = "Column"
            else:
                element_type = "Unbekannter Typ"

            # Aktualisieren des Namens des IFC-Elements
            ifc_element.Name = new_name

            # Aktualisieren des Materials des IFC-Elements
            update_ifc_material(ifc_element, new_material)

            print(f"Element mit GUID {guid} gefunden. Typ: {element_type}, Neuer Name: {new_name}")

# Funktion zum Aktualisieren des IFC-Materials
def update_ifc_material(ifc_element, new_material):
    # Extrahiere das aktuelle Material des IFC-Elements
    current_material = ifcopenshell.util.element.get_material(ifc_element)

    # Prüfe, ob das Material existiert und aktualisiere es
    if current_material:
        # Setze das neue Material
        current_material.Name = new_material

# Speichern der Änderungen in der IFC-Datei
ifc_file.write(ifc_file_path)

# Schließen der IFC-Datei
#ifc_file.close()

# Schließen der Excel-Datei
workbook.close()

print("Die IFC-Datei wurde aktualisiert und gespeichert.")

