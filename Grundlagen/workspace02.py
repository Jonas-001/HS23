import ifcopenshell
import openpyxl

# Dateipfade zu IFC- und Excel-Dateien
ifc_file_path = r"C:\Users\joni_\Documents\Studium HSLU\VS_Projects\IFC-Editor\ARC_Modell_NEST_230328.ifc"
excel_file_path = r"C:\Users\joni_\Documents\Studium HSLU\VS_Projects\IFC-Editor\Export_Daten_IFC - Test.xlsx"

# Öffnen der IFC-Datei
ifc_file = ifcopenshell.open(ifc_file_path)

# Öffnen der Excel-Datei
workbook = openpyxl.load_workbook(excel_file_path)

# Iteration durch alle Sheets in der Excel-Datei
for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]

    # Iteration durch die Zeilen in der Excel-Datei
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        guid = row[0]  # Die GUID-ID befindet sich in der ersten Spalte
        element_typ_name = row[2]  # Der Typname des IFC-Elements befindet sich in der dritten Spalte
        eigenschaft_name = row[3]  # Der Name der Eigenschaft des IFC-Elements befindet sich in der vierten Spalte

        # Suchen des entsprechenden IFC-Elements anhand der GUID-ID
        ifc_element = ifc_file.by_guid(guid)
    
        # Überprüfen, ob das IFC-Element existiert
        if ifc_element:
            # Zugriff auf die gewünschte Eigenschaft des IFC-Elements
            try:
                wert = getattr(getattr(ifc_element, element_typ_name), eigenschaft_name)
                
                print(f"Die Eigenschaft {eigenschaft_name} für das Element mit der GUID {guid} und Typ {element_typ_name} wurde gefunden.")
                
                # Du kannst hier den Wert weiterverarbeiten oder Änderungen vornehmen, falls erforderlich
            except AttributeError:
                print(f"Die Eigenschaft {eigenschaft_name} oder der Typ {element_typ_name} wurde nicht gefunden.")
            
# Schliessen der IFC-Datei
#ifc_file.close()

# Schliessen der Excel-Datei
workbook.close()


